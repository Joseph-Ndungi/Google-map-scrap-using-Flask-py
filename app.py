from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.common.exceptions import NoSuchElementException
import time 
import string
import openpyxl
import os
from selenium.webdriver.firefox.service import Service


s=Service(r'D:\my projects\geckodriver.exe')
driver = webdriver.Firefox(service=s)
wait = WebDriverWait(driver, 5)

#Opening Google maps 
url='https://www.google.com/maps'
driver.get(url)

#Finding the search box
searchbox=driver.find_element(By.ID,'searchboxinput')
location= "garage door repair provo utah"
searchbox.send_keys(location)
searchbox.send_keys(Keys.ENTER)
time.sleep(2)

#Locating the results
entries=driver.find_elements(By.CLASS_NAME,'m6QErb DxyBCb kA9KIf dS8AEf ecceSd')


#Prepare the excel file using the Openpyxl  
wb= openpyxl.load_workbook(r'D:\my projects\Google-map-scrap-using-Flask-py\garages.xlsx')
sheetname=wb.sheetnames
sheet=wb[sheetname[0]]
# sheet.cell(row=1,column=1).value="Name"
# sheet.cell(row=1,column=2).value="Address"
# sheet.cell(row=1,column=3).value="Phone"
sheet.title ="garages"


print(entries)
#Extracting the information from the results  
for entry in entries:
    #Empty list 
    labels=[]
    #Extracting the Name, adress, Phone, and website:
    
    name= entry.get_attribute("aria-label")
    adress= entry.find_element_by_class_name('section-result-location').text
    phone = entry.find_element_by_class_name('section-result-hours-phone-container').text.split(' · ')[-1]
    try:
        webcontainer= entry.find_element_by_class_name('section-result-action-container')
        website=entry.find_element_by_tag_name('a').get_attribute("href")
        print(name)
        print(adress)
    except NoSuchElementException:
        website="No website could be found"
        
       

    #Try/except  to write the extracted info in the Excel file pass if doessn't exist 
    try:
        sheet.append([location,name,adress,phone,website])
    except IndexError:
        pass
 
#saving the excel file 
wb.save("companies.xlsx")